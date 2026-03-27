import os
import uuid
from datetime import datetime, timezone
from flask import Flask, request, jsonify
from openpyxl import load_workbook, Workbook

app = Flask(__name__)

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data_0")
os.makedirs(DATA_DIR, exist_ok=True)


def get_account_file(account_id: str) -> str:
    """Get the xlsx file path for an account."""
    return os.path.join(DATA_DIR, f"{account_id}.xlsx")


def init_account_file(filepath: str) -> Workbook:
    """Initialize a new xlsx file with headers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Memos"
    ws.append(["id", "title", "content", "from", "created_at", "updated_at"])
    wb.save(filepath)
    return wb


def load_account_wb(account_id: str) -> Workbook:
    """Load workbook for account, create if not exists."""
    filepath = get_account_file(account_id)
    if not os.path.exists(filepath):
        return init_account_file(filepath)
    return load_workbook(filepath)


def row_to_memo(row: tuple) -> dict:
    """Convert xlsx row to memo dict."""
    if len(row) == 5:
        # Old format: id, title, content, created_at, updated_at
        return {
            "id": row[0],
            "title": row[1],
            "content": row[2],
            "from": None,
            "created_at": row[3],
            "updated_at": row[4],
        }
    # New format: id, title, content, from, created_at, updated_at
    return {
        "id": row[0],
        "title": row[1],
        "content": row[2],
        "from": row[3] if len(row) > 3 else None,
        "created_at": row[4] if len(row) > 4 else row[3],
        "updated_at": row[5] if len(row) > 5 else row[4],
    }


def memo_to_row(memo: dict) -> list:
    """Convert memo dict to xlsx row."""
    return [
        memo["id"],
        memo["title"],
        memo["content"],
        memo.get("from"),
        memo["created_at"],
        memo["updated_at"],
    ]


def find_memo_row(ws, memo_id: str) -> int | None:
    """Find row index for memo by id, return 1-based index or None."""
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == memo_id:
            return i
    return None


@app.route("/api/accounts/<account_id>/memos", methods=["POST"])
def create_memo(account_id: str):
    """Create a new memo, optionally sending to another user."""
    data = request.get_json()
    if not data or not data.get("title"):
        return jsonify({"success": False, "error": "title is required"}), 400

    now = datetime.now(timezone.utc).isoformat()
    title = data["title"][:200]
    to_user = data.get("to")

    # Cross-user memo: prefix title and optionally store in recipient's file
    if to_user:
        title = f"[给{to_user}] {title}"
        # Store in recipient's file instead of sender's
        target_account = to_user
        memo = {
            "id": str(uuid.uuid4()),
            "title": title,
            "content": data.get("content", ""),
            "from": account_id,
            "created_at": now,
            "updated_at": now,
        }
    else:
        target_account = account_id
        memo = {
            "id": str(uuid.uuid4()),
            "title": title,
            "content": data.get("content", ""),
            "created_at": now,
            "updated_at": now,
        }

    wb = load_account_wb(target_account)
    ws = wb["Memos"]
    ws.append(memo_to_row(memo))
    wb.save(get_account_file(target_account))

    return jsonify({"success": True, "data": memo}), 201


@app.route("/api/accounts/<account_id>/memos", methods=["GET"])
def list_memos(account_id: str):
    """List all memos for an account."""
    if not os.path.exists(get_account_file(account_id)):
        return jsonify({"success": True, "data": []})

    wb = load_workbook(get_account_file(account_id))
    ws = wb["Memos"]
    memos = [row_to_memo(row) for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
    return jsonify({"success": True, "data": memos})


@app.route("/api/accounts/<account_id>/memos/<memo_id>", methods=["GET"])
def get_memo(account_id: str, memo_id: str):
    """Get a single memo."""
    if not os.path.exists(get_account_file(account_id)):
        return jsonify({"success": False, "error": "Account not found"}), 404

    wb = load_workbook(get_account_file(account_id))
    ws = wb["Memos"]
    row_idx = find_memo_row(ws, memo_id)

    if row_idx is None:
        return jsonify({"success": False, "error": "Memo not found"}), 404

    row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
    return jsonify({"success": True, "data": row_to_memo(row)})


@app.route("/api/accounts/<account_id>/memos/<memo_id>", methods=["PUT"])
def update_memo(account_id: str, memo_id: str):
    """Update a memo."""
    if not os.path.exists(get_account_file(account_id)):
        return jsonify({"success": False, "error": "Account not found"}), 404

    data = request.get_json()
    if not data:
        return jsonify({"success": False, "error": "Request body required"}), 400

    wb = load_workbook(get_account_file(account_id))
    ws = wb["Memos"]
    row_idx = find_memo_row(ws, memo_id)

    if row_idx is None:
        return jsonify({"success": False, "error": "Memo not found"}), 404

    row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
    memo = row_to_memo(row)

    if "title" in data:
        memo["title"] = data["title"][:200]
    if "content" in data:
        memo["content"] = data["content"]

    memo["updated_at"] = datetime.now(timezone.utc).isoformat()

    for col, val in enumerate(memo_to_row(memo), start=1):
        ws.cell(row=row_idx, column=col, value=val)

    wb.save(get_account_file(account_id))
    return jsonify({"success": True, "data": memo})


@app.route("/api/accounts/<account_id>/memos/<memo_id>", methods=["DELETE"])
def delete_memo(account_id: str, memo_id: str):
    """Delete a memo."""
    if not os.path.exists(get_account_file(account_id)):
        return jsonify({"success": False, "error": "Account not found"}), 404

    wb = load_workbook(get_account_file(account_id))
    ws = wb["Memos"]
    row_idx = find_memo_row(ws, memo_id)

    if row_idx is None:
        return jsonify({"success": False, "error": "Memo not found"}), 404

    ws.delete_rows(row_idx)
    wb.save(get_account_file(account_id))
    return jsonify({"success": True, "data": None})


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5001, debug=True)
